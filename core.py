import pandas as pd
import pdfplumber
import re
from openpyxl import load_workbook
from io import BytesIO
from pdf2image import convert_from_bytes
import pytesseract
from PIL import Image
import streamlit as st

def extract_text_from_pdf(file):
    text = ""
    try:
        with pdfplumber.open(file) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"

        if not text.strip():
            st.write("📷 Переход на OCR: текст не найден через pdfplumber")
            file.seek(0)
            images = convert_from_bytes(file.read(), dpi=300)
            ocr_text = ""
            for img in images:
                ocr_text += pytesseract.image_to_string(img, lang="rus") + "\n"
            if not ocr_text.strip():
                raise ValueError("Не удалось распознать текст даже через OCR.")
            text = ocr_text

    except Exception as e:
        raise ValueError(f"Ошибка при извлечении текста из PDF: {e}")

    return text

def parse_requirements(text):
    lines = text.split("\n")
    rows = []
    for line in lines:
        if any(char.isdigit() for char in line):
            parts = re.split(r"\s{2,}|\t", line.strip())
            if len(parts) >= 2:
                name = parts[0]
                quantity = re.search(r"\d+", parts[1])
                quantity = quantity.group() if quantity else ""
                rows.append({"Наименование из ТЗ": name, "Кол-во": quantity})
    st.write(f"📄 Распознано строк в ТЗ: {len(rows)}")
    return pd.DataFrame(rows)

def load_price_list(files):
    all_items = []
    for file in files:
        df = pd.read_excel(file, header=None)
        for index, row in df.iterrows():
            for col in row:
                if isinstance(col, str) and any(keyword in col.lower() for keyword in ["стол", "кресло", "лампа", "шкаф", "банкетка", "барьер"]):
                    item = {
                        "Артикул": row[0] if len(row) > 1 else "",
                        "Наименование": col,
                        "Цена": next((v for v in row if isinstance(v, (int, float))), "")
                    }
                    all_items.append(item)
                    break
    st.write(f"📊 Загружено позиций из прайсов: {len(all_items)}")
    return pd.DataFrame(all_items)

def load_discounts(file):
    df = pd.read_excel(file)
    discounts = dict(zip(df.iloc[:, 0], df.iloc[:, 1]))
    st.write(f"💸 Считано скидок: {len(discounts)}")
    return discounts

def process_documents(spec_file, prices_files, discounts_file=None):
    st.write("🔍 Чтение и распознавание ТЗ...")
    text = extract_text_from_pdf(spec_file)
    ts_text = text[:1000] if text else "(Текст не найден)"

    st.write("📑 Парсинг строк из ТЗ...")
    spec_df = parse_requirements(text)

    st.write("📥 Загрузка прайсов...")
    prices_df = load_price_list(prices_files)

    discounts = {}
    if discounts_file:
        st.write("💼 Загрузка скидок...")
        discounts = load_discounts(discounts_file)

    st.write("🔄 Сопоставление позиций...")
    results = []
    for _, row in spec_df.iterrows():
        name = row["Наименование из ТЗ"]
        qty = row["Кол-во"]
        matches = prices_df[prices_df["Наименование"].str.contains(name.split()[0], case=False, na=False)].head(3)

        item = {
            "Наименование из ТЗ": name,
            "Кол-во": qty
        }

        for i, (_, match_row) in enumerate(matches.iterrows()):
            price = match_row.get("Цена")
            supplier = match_row.get("Поставщик", f"Поставщик {i+1}")
            discount = discounts.get(supplier, 0)
            final_price = round(price * (1 - discount / 100), 2) if price else ""

            item[f"Поставщик {i+1}"] = supplier
            item[f"Цена {i+1}"] = price
            item[f"Скидка {i+1}"] = f"{discount}%"
            item[f"Итог {i+1}"] = final_price

        results.append(item)

    result_df = pd.DataFrame(results)
    st.write(f"✅ Готово! Совпадений: {len(result_df)} строк")

    st.write("📤 Формирование Excel-файла...")
    template = "Форма для результата.xlsx"
    wb = load_workbook(template)
    ws = wb.active

    start_row = 10
    for i, row in result_df.iterrows():
        ws.cell(row=start_row + i, column=1, value=i + 1)
        ws.cell(row=start_row + i, column=2, value=row["Наименование из ТЗ"])
        ws.cell(row=start_row + i, column=3, value=row["Кол-во"])
        ws.cell(row=start_row + i, column=4, value=row.get("Поставщик 1"))
        ws.cell(row=start_row + i, column=5, value=row.get("Цена 1"))
        ws.cell(row=start_row + i, column=6, value=row.get("Скидка 1"))
        ws.cell(row=start_row + i, column=7, value=row.get("Итог 1"))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return ts_text, result_df, output.read()
